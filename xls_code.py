import openpyxl as op



class dataupdating:



    def __init__(self,file_name:str,sheet_name):



        self.excel_file_name = file_name
        self. work = op.load_workbook(filename=r'{0}'.format(self.excel_file_name))
       
        self.sheet = self.work[sheet_name]
        self.row = self.sheet.max_row
        self.colum = self.sheet.max_column


    def creating_sheet_name(self,sheet_name:str):
        self.work.create_sheet(sheet_name)
        self.work.save(self.excel_file_name)



       
    
    def creating_headders(self,colum_name1:str,colum_name2:str):
        '''
        headder updating 
        '''
        for a in range(1,2):
            for index ,data  in enumerate([colum_name1,colum_name2],start=1):
                self.sheet.cell(a,index).value = data

        self.work.save(self.excel_file_name)


    def data_updating(self,cell_data1,cell_data2):
        '''cell dataupdating...'''
        for a in range(self.row,self.row+1):
            for index ,data  in enumerate([cell_data1,cell_data2],start=1):
                self.sheet.cell(a,index).value = data
        self.work.save(self.excel_file_name)